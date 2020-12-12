import  openpyxl
from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference, Series, LineChart, ScatterChart
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font

thick_border = Border(left=Side(style='thick'),
                     right=Side(style='thick'),
                     top=Side(style='thick'),
                     bottom=Side(style='thick'))

Seç = ""
while Seç != "1" and Seç != "2":
    print("1- Öğrencileri yerleştir\n2- Hesaplama yap\n")
    Seç = input("Giriş:")
    print("\n---------------------------------------------------------------------\n")

if Seç == "1":
    PATH = "CTR.xlsx"
    excel = openpyxl.load_workbook(PATH)
    ws = excel.worksheets[0]
    Öğrenciler = []
    Konular = []
    a = ""
    print("1- Bütün Öğrencileri gör\n2- Son öğrenciyi sil\n3- Devam")
    while a != "3":
        a = input("Öğrenci ismi:")
        if a == "1":
            yazı = ""
            for i in range(len(Öğrenciler)):
                if i % 3 == 0 and i != 0:
                    yazı += "\n"
                yazı += Öğrenciler[i] + "\t"
            print("---------------Öğrenciler----------------")
            print(yazı)
            print("-----------------------------------------")
        elif a == "2":
            if len(Öğrenciler) > 0:
                print("------Son öğrenci başarı ile silindi----------")
                Öğrenciler.pop(-1)
            else:
                print("------Henüz öğrenci girmediniz------")
        elif a != "3":
            Öğrenciler.append(a)
    a = ""
    print("\n1- Bütün konuları gör\n2- Son konuyu sil\n3- Devam")
    while a != "3":
        a = input("Konu ismi:")
        if a == "1":
            yazı = ""
            for i in range(len(Konular)):
                if i % 3 == 0:
                    yazı += "\n"
                yazı += Konular[i] + "\t"
            print("---------------Konular----------------")
            print(yazı)
            print("-----------------------------------------")
        elif a == "2":
            if len(Konular) > 0:
                print("------Son Konu başarı ile silindi----------")
                Konular.pop(-1)
            else:
                print("------Henüz konu girmediniz------")
        elif a != "3":
            Konular.append(a)
    x = 1
    y = 1

    for a in Öğrenciler:
        ws.cell(y, x).value = a
        ws.cell(y, x).border = thick_border
        for b in Konular:
            y += 1
            ws.cell(y, x).value = b
            for c in range(x, x + 11):
                ws.cell(y, c).border = thick_border
        y += 2
    excel.worksheets[0].title = "Liste"
    try:
        excel.save("Liste.xlsx")
    except PermissionError:
        print("Lütfen excel dosyasını kapatınca devam yazıp enter tuşuna basın")
        input("Giriş:")
        excel.save("Liste.xlsx")
    print("------------------------------------------------\nİŞLEMİNİZ BAŞARI İLE GERÇEKLEŞTİ\n------------------------------------------------")
elif Seç == "2":
    try:
        PATH = "Liste.xlsx"
        excel = openpyxl.load_workbook(PATH)
    except:
        print("Görünüşe bakılırsa Liste isimli excel dosyasının ismini değiştirmişsiniz!\nlütfen yeni ismini girin")
        PATH = input("yeni isim:") + ".xlsx"
        excel = openpyxl.load_workbook(PATH)

    ws = excel.worksheets[0]
    Öğrenciler = []
    Konular = []
    x = 1
    y = 1

    Öğrenciler.append(ws.cell(y,x).value)
    y += 1
    while ws.cell(y,1).value != None:
        Konular.append(ws.cell(y,1).value)
        y += 1
    y += 1
    while (ws.cell(y,1).value != None or ws.cell(y+1,1).value != None) == True:
        Öğrenciler.append(ws.cell(y,1).value)
        y += len(Konular) +2
    TamListe = []
    y = 1
    x = 1
    kök = y

    while y < len(Öğrenciler) * (len(Konular)+2):
        kök = y
        new = [ws.cell(y,1).value]
        y += 1
        while y < kök + len(Konular) +1:
            toplam = 0
            Sayı = 0
            x = 2
            while ws.cell(y,x).value != None:
                toplam += int(ws.cell(y,x).value)
                Sayı += 1
                x += 1
            if Sayı != 0:
                new.append(round(toplam / Sayı, 1))
            y += 1
        y += 1
        TamListe.append(new)
    
    Sheet = excel.worksheets[1]
    y = 1
    x = 1
    for i in range(len(Konular)):
        y = 1
        Sheet.cell(y, x+1).value = "Kişi"
        Sheet.cell(y, x+2).value = "Sınıf Ort."
        ort = 0
        for c in range(len(TamListe)):

            ort += TamListe[c][i+1]
        ort /= len(TamListe)
        ort = round(ort,1)
        y = 2
        for b in range(len(TamListe)):
          Sheet.cell(y,x).value = TamListe[b][0]
          Sheet.cell(y,x+1).value = TamListe[b][i+1]
          Sheet.cell(y, x + 2).value = ort
          y += 1

        x += 3
    Pos = "A15"
    number = 1
    numb2 = 15
    Alfa = [[1, "A"], [2, "B"], [3, "C"], [4, "D"], [5, "E"], [6, "F"], [7, "G"], [8, "H"], [9, "I"], [10, "J"],
            [11, "K"], [12, "L"], [13, "M"], [14, "N"], [15, "O"], [16, "P"], [17, "Q"], [18, "R"], [19, "S"],
            [20, "T"], [21, "U"], [22, "V"], [23, "W"], [24, "X"], [25, "Y"], [26, "Z"]]

    #------------------------bar grafiği---------------------------------
    x = 1
    y = 1
    for i in range(len(Konular)):
        chart1 = BarChart()
        chart1.type = "col"
        chart1.style = 10
        chart1.title = Konular[i]
        chart1.y_axis.title = ''
        chart1.x_axis.title = ''

        data = Reference(Sheet, min_col=x+1, min_row=1, max_row=len(TamListe) + 1, max_col=x+2)
        cats = Reference(Sheet, min_col=1, min_row=2, max_row=len(TamListe) + 1)
        chart1.add_data(data, titles_from_data=True)
        chart1.set_categories(cats)
        chart1.shape = 4
        Sheet.add_chart(chart1, Pos)

        if number + 10 < 26:
            number += 10
        else:
            number = 1
            numb2 += 15
        Pos = str(Alfa[number - 1][1]) + str(numb2)
        x += 3
    excel.worksheets[1].title = "Grafikler"
    excel.worksheets[0].title = "Liste"

    excel.save(PATH)
    print(
        "------------------------------------------------\nİŞLEMİNİZ BAŞARI İLE GERÇEKLEŞTİ\n------------------------------------------------")





